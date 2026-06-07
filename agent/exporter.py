"""
Excel导出器 — 生成结构化岗位汇总表
"""
import os
from typing import List
from .models import JobPosition, SearchResult

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """将搜索结果导出为格式化的Excel文件"""

    # 列定义
    COLUMNS = [
        ("序号", 6),
        ("公司名称", 18),
        ("岗位名称", 22),
        ("岗位方向", 12),
        ("行业", 12),
        ("工作地点", 10),
        ("实习薪资", 14),
        ("岗位职责概述", 40),
        ("任职要求概述", 40),
        ("投递链接/来源", 35),
        ("备注", 25),
    ]

    # 样式定义
    HEADER_FILL = PatternFill(start_color="0052D9", end_color="0052D9", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    CONTENT_FONT = Font(name="微软雅黑", size=10)
    LINK_FONT = Font(name="微软雅黑", size=10, color="0052D9", underline="single")
    CONVERSION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def __init__(self, result: SearchResult):
        self.result = result
        self.result.sort_by_priority()

    def export(self, output_path: str) -> str:
        """导出到Excel文件"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "岗位汇总"

        # 写入表头
        self._write_header(ws)

        # 写入数据
        for i, pos in enumerate(self.result.positions, 1):
            self._write_row(ws, i + 1, i, pos)

        # 设置列宽
        self._set_column_widths(ws)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}{len(self.result.positions) + 1}"

        # 保存
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        wb.save(output_path)
        return output_path

    def _write_header(self, ws):
        """写入表头"""
        for col_idx, (col_name, _) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _write_row(self, ws, row_idx: int, seq: int, pos: JobPosition):
        """写入一行数据"""
        data = pos.to_dict()
        values = [
            seq,                          # 序号
            data["公司名称"],
            data["岗位名称"],
            data["岗位方向"],
            data["行业"],
            data["工作地点"],
            data["实习薪资"],
            data["岗位职责概述"],
            data["任职要求概述"],
            data["投递链接/来源"],
            data["备注"],
        ]

        is_conversion = pos.has_conversion

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = self.CONTENT_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER

            # 有转正机会的行标记绿色背景
            if is_conversion:
                cell.fill = self.CONVERSION_FILL

        ws.row_dimensions[row_idx].height = max(28, 15 * (len(pos.responsibilities) // 60 + 1))


    def _set_column_widths(self, ws):
        """设置列宽"""
        for col_idx, (_, width) in enumerate(self.COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def export_from_json(json_path: str, output_path: str) -> str:
    """从JSON文件导入并导出Excel（命令行接口）"""
    import json
    with open(json_path, "r", encoding="utf-8") as f:
        positions_data = json.load(f)

    positions = []
    for item in positions_data:
        pos = JobPosition(
            company=item.get("公司名称", ""),
            title=item.get("岗位名称", ""),
            direction=item.get("岗位方向", "数据分析"),
            industry=item.get("行业", "其他"),
            city=item.get("工作地点", ""),
            salary=item.get("实习薪资", "面议"),
            responsibilities=item.get("岗位职责概述", ""),
            requirements=item.get("任职要求概述", ""),
            apply_link=item.get("投递链接/来源", ""),
            source=item.get("来源平台", ""),
            has_conversion=item.get("有转正机会", False),
            can_retain=item.get("可留用", False),
            target_grad=item.get("面向届别", ""),
            notes=item.get("备注", ""),
        )
        positions.append(pos)

    result = SearchResult(positions=positions, total_found=len(positions))
    exporter = ExcelExporter(result)
    return exporter.export(output_path)
